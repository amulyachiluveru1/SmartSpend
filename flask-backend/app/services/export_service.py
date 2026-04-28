import csv
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def generate_csv(expenses):
    """Generate CSV data from expense list."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date', 'Category', 'Amount', 'Description'])
    for exp in expenses:
        writer.writerow([
            exp.date.isoformat(),
            exp.category.name,
            float(exp.amount),
            exp.description or ''
        ])
    output.seek(0)
    return output.getvalue().encode('utf-8')

def generate_excel(expenses):
    """Generate Excel file from expense list with proper formatting."""
    # Create a list of dictionaries for pandas
    data = []
    for exp in expenses:
        data.append({
            'Date': exp.date.strftime('%Y-%m-%d'),
            'Category': exp.category.name,
            'Amount': float(exp.amount),
            'Description': exp.description or ''
        })
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    # Write headers
    headers = ['Date', 'Category', 'Amount', 'Description']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        # Style headers
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=row_data['Date'])
        ws.cell(row=row_idx, column=2, value=row_data['Category'])
        ws.cell(row=row_idx, column=3, value=row_data['Amount'])
        ws.cell(row=row_idx, column=4, value=row_data['Description'])
        
        # Format amount as currency
        ws.cell(row=row_idx, column=3).number_format = '$#,##0.00'
    
    # Add summary section
    summary_row = len(data) + 3
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
    
    # Calculate totals
    total_amount = sum(item['Amount'] for item in data)
    avg_amount = total_amount / len(data) if data else 0
    unique_categories = len(set(item['Category'] for item in data))
    
    summary_data = [
        ("Total Expenses:", f"${total_amount:,.2f}"),
        ("Average Amount:", f"${avg_amount:,.2f}"),
        ("Categories Used:", str(unique_categories)),
        ("Total Records:", str(len(data))),
        ("Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    ]
    
    for idx, (label, value) in enumerate(summary_data, 1):
        ws.cell(row=summary_row + idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=summary_row + idx, column=2, value=value)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    return excel_bytes.getvalue()